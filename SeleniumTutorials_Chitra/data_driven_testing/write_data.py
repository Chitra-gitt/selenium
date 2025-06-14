'''
Created on 11 Jun 2025

@author: Chitra
'''
import openpyxl
filepath = "/Users/chitra/Desktop/Testdata-DDT.xlsx" 
my_workbook = openpyxl.load_workbook(filepath)
my_active_sheet = my_workbook["Sheet1"]

my_active_sheet.cell(2, 4).value = "Pass"
my_workbook.save(filepath)  

my_active_sheet.cell(3, 4).value = "Fail"
my_workbook.save(filepath)  

my_active_sheet.cell(4, 4).value = "Fail"
my_workbook.save(filepath)  

my_active_sheet.cell(5, 4).value = "Fail"
my_workbook.save(filepath)  

my_active_sheet.cell(6, 4).value = "Pass"
my_workbook.save(filepath)  

