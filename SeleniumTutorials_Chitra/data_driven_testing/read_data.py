'''
Created on 8 Jun 2025

@author: Chitra
'''

import openpyxl

filepath = "/Users/chitra/Desktop/Testdata-DDT.xlsx" 
my_workbook = openpyxl.load_workbook(filepath)
my_active_sheet = my_workbook.active
total_rows = my_active_sheet.max_row
print(total_rows)


total_columns = my_active_sheet.max_column
print(total_columns)


'''
o/p- 

12
4

'''

import openpyxl

filepath = "/Users/chitra/Desktop/Testdata-DDT.xlsx" 
my_workbook = openpyxl.load_workbook(filepath)
my_active_sheet = my_workbook.active # to get the currently active sheet
my_active_sheet = my_workbook["Sheet1"]
total_rows = my_active_sheet.max_row
print(total_rows)


total_columns = my_active_sheet.max_column
print(total_columns)

for i in range(2, total_rows+1):
    username = my_active_sheet.cell(i, 1).value
    password = my_active_sheet.cell(i, 2).value
    print(username, password)








